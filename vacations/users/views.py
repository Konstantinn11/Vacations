from django.urls import reverse, reverse_lazy
from django.contrib.auth import get_user_model
User = get_user_model()
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from .models import User_info, Unit, Vacation, Position, Tag
from .forms import VacationForm, TagForm, UnitForm
import datetime as dt
import pandas as pd
from django.contrib import messages as msg
from calendar import monthrange
import json
from .vacation_data import months_ru, color_cycle
from holiday_calendar.utils import month_num_str
from holiday_calendar.models import Holiday
from django.http import JsonResponse
from django.contrib.auth.views import PasswordResetDoneView, PasswordResetView
from django.contrib.auth.forms import PasswordResetForm
from django.http import HttpResponse
import os
from openpyxl import load_workbook
import urllib.parse
import copy
import re
import pymorphy2
from openpyxl.styles import Alignment, Border, Side, Font
from holiday_calendar.utils import get_calendar_data
from .utils import calculate_end_date, calculate_working_days, get_bosses_dict
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db.models import Count, Q
from django.core.mail import send_mail
from django.db import IntegrityError, transaction

class CustomPasswordResetView(PasswordResetView):
    form_class = PasswordResetForm
    success_url = reverse_lazy('password_reset_done')

    def form_valid(self, form):
        # Получаем введенный email
        email = form.cleaned_data.get('email')

        # Проверяем, существует ли пользователь с таким email
        if not User.objects.filter(email=email).exists():
            msg.error(self.request, "Такого адреса электронной почты не существует")
            return self.form_invalid(form)

        # Сохраняем email в сессии
        self.request.session['email'] = email

        return super().form_valid(form)


class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'registration/password_reset_done.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Получаем email из сессии
        email = self.request.session.get('email')
        if email:
            context['email'] = email
        return context
    

def home_view(request):
    return redirect(reverse('vac_all', kwargs={'otd': 0}))


def server_error(request):
    return render(request, "misc/500.html", status=500)


def page_not_found(request, exception):
    # Переменная exception содержит отладочную информацию,
    return render(request, "misc/404.html",{"path": request.path},status=404)


def get_key_from_dict_by_value(dict, value):
    return [k for k, v in dict.items() if v == value][0]


def get_cross_vacations(vacations, users_colors, month_num_str):
    # Построим граф пересечений
    n = len(vacations)
    adj = {i: set() for i in range(n)}
    
    def is_overlap(a, b):
        # Возвращает True, если интервалы [a.day_start, a.day_end] и [b.day_start, b.day_end] пересекаются
        return not (a.day_end < b.day_start or b.day_end < a.day_start)
    
    for i in range(n):
        for j in range(i + 1, n):
            if is_overlap(vacations[i], vacations[j]):
                adj[i].add(j)
                adj[j].add(i)
    
    # Найдём все компоненты связности в этом графе
    visited = set()
    groups = []
    
    for i in range(n):
        if i not in visited:
            stack = [i]
            comp = []
            visited.add(i)
            while stack:
                u = stack.pop()
                comp.append(u)
                for v in adj[u]:
                    if v not in visited:
                        visited.add(v)
                        stack.append(v)

            if len(comp) > 1:
                groups.append(comp)
    
    # Для каждой группы компонент сконструируем список словарей с данными и цветом
    result = []
    for comp in groups:
        vac_list = []
        for idx in comp:
            vac = vacations[idx]
            user_id = vac.user.id
            vac_list.append({
                'vac': vac,
                'range': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} – "
                         f"{vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
                'color': users_colors.get(user_id, "#cccccc"),
            })
        vac_list.sort(key=lambda x: x['vac'].day_start)
        result.append(vac_list)

    return result


def copy_dict_for_js(month_all):
    month_all_for_js = {}
    for month, days in month_all.items():
        month_all_for_js[month] = {}
        for week, days_in_week in days.items():
            month_all_for_js[month][week] = {}
            for i in range(len(days_in_week)):
                month_all_for_js[month][week][i] = {}
                month_all_for_js[month][week][i]['name'] = month_all[month][week][i]['name']
                month_all_for_js[month][week][i]['date'] = {}
                for key in month_all[month][week][i]['data'].keys():
                    # Преобразование ключа в строку, если он не строка
                    key_str = str(key).replace(' ', '_')
                    month_all_for_js[month][week][i]['date'][key_str] = month_all[month][week][i]['data'][key]
                month_all_for_js[month][week][i]['data'] = month_all[month][week][i]['data']
    return month_all_for_js


def full_year(year):
    month_all = {
        1:'Январь', 2:'Февраль', 3:'Март',
        4:'Апрель', 5:'Май', 6:'Июнь',
        7:'Июль', 8:'Август', 9:'Сентябрь',
        10:'Октябрь', 11:'Ноябрь', 12:'Декабрь',
        }
    month_new = {}
    for i in range(1, 13):
        day = dt.datetime.today().replace(
            year=year,
            month=i,
            day=1
            )
        mnth_strt_d = day.replace(month=i, day=1).weekday()
        days_in_month = monthrange(year, i)[1]
        weeks, k = {}, 1
        weeks[k] = []
        # Добавляем пустые клетки в начале месяца, если он начался не с понедельника
        [weeks[k].append('') for j in range(mnth_strt_d) if mnth_strt_d != 0]
        # Заполняем месяц
        for j in range(1, days_in_month + 1):
            if mnth_strt_d < 7:
                weeks[k].append(j)
                mnth_strt_d += 1
            else:
                k += 1
                weeks[k] = []
                weeks[k].append(j)
                mnth_strt_d = 1
        # Добавляем пустые клетки в конце месяца, если он закончился не в воскресенье    
        [weeks[k].append('') for j in range(mnth_strt_d, 7)]
        month_new[month_all[i]] = weeks
    return month_new


def vac_all(request, otd):
    today = dt.datetime.today().date()
    year = today.year
    holidays, _ = get_calendar_data(year)
    bosses = get_bosses_dict()

    current_user_name = request.user.get_full_name()

    # Если пользователь не является боссом
    if current_user_name not in bosses:
        otd_id = User_info.objects.filter(user_id=request.user.id)[0].otd_number_id
        unit = Unit.objects.filter(id=otd_id).first()
        if unit:
            otd = int(unit.description)
        else:
            otd = 0
        otd_users = User_info.objects.filter(otd_number_id=otd_id)
        otd_users_id = [user.user_id for user in otd_users]
        vacations = Vacation.objects.filter(user_id__in=otd_users_id, day_end__gte=today).order_by('day_start')
        otds_for_choise = [otd]
    else:
        # Если пользователь босс
        if otd == 0:  # Все отделы, которыми он руководит
            otds_for_choise = bosses[current_user_name]
            otd_ids = [Unit.objects.filter(description=descr).first().id for descr in otds_for_choise if Unit.objects.filter(description=descr).exists()]
            otd_users = User_info.objects.filter(otd_number_id__in=otd_ids)
        else:  # Выбран конкретный отдел
            unit = Unit.objects.filter(description=otd).first()
            if unit:
                otd_id = unit.id
                otd_users = User_info.objects.filter(otd_number_id=otd_id)
                otds_for_choise = [otd]
            else:
                otd_id = None
                otd_users = User_info.objects.none()
                otds_for_choise = []

        otd_users_id = [user.user_id for user in otd_users]
        vacations = Vacation.objects.filter(user_id__in=otd_users_id, day_end__gte=today).order_by('day_start')

    # Находим ближайшие отпуска
    nearest_vacations = {}
    for vac in vacations:
        if vac.day_start >= today or (vac.day_start <= today <= vac.day_end):
            if vac.user.id not in nearest_vacations or vac.day_start < nearest_vacations[vac.user.id].day_start:
                nearest_vacations[vac.user.id] = vac

    filtered_vacations = list(nearest_vacations.values())

    vacation_start_dates = {}
    for vac in filtered_vacations:
        vacation_start_dates.setdefault(vac.user.get_full_name(), []).append(vac.day_start)

    users_colors = {}
    for vac in vacations:
        if vac.user.id not in users_colors:
            users_colors[vac.user.id] = next(color_cycle)

    # Подготовка данных по пользователям
    vacations_by_user = {}
    users_otd = User_info.objects.all()
    for vac in filtered_vacations:
        uid = vac.user.id
        if uid not in vacations_by_user:
            user_info = users_otd.get(user_id=uid)
            position = user_info.position.position if user_info.position else "Не указана"
            vacations_by_user[uid] = {
                'name': vac.user.get_full_name(),
                'user_id': uid,
                'position': position,
                'otd': otd,
                'color': users_colors[uid],
                'sum': 0,
                'in_vacation': vac.day_start <= today <= vac.day_end,
                'dates': [],
                'vacation_periods': [],
                'vacation_start_dates': []
            }
            # Обновляем отдел
            for u in users_otd:
                if u.user_id == uid:
                    vacations_by_user[uid]['otd'] = u.otd_number
                    break

        days_count = vac.how_long

        # Обновляем данные пользователя
        vacations_by_user[uid]['vacation_start_dates'].append((vac.day_start, days_count))
        vacations_by_user[uid]['sum'] += days_count
        vacations_by_user[uid]['dates'].append({
            'd': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} - {vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
            'vac_id': vac.id,
        })
        formatted_start = f"{vac.day_start.day} {months_ru[vac.day_start.month]} {vac.day_start.year}"
        formatted_end = f"{vac.day_end.day} {months_ru[vac.day_end.month]} {vac.day_end.year}"
        vacations_by_user[uid]['vacation_periods'].append(f"{formatted_start} - {formatted_end}")

    # Сортировка
    sort_otd    = request.GET.get('sort_otd')              
    sort_period = request.GET.get('sort_period') or 'asc'

    vacations_by_user_items = list(vacations_by_user.items())

    for key, direction in request.GET.items():
        if key == 'sort_otd' and direction in ('asc','desc'):
            rev = (direction == 'desc')
            vacations_by_user_items.sort(
                key=lambda item: str(item[1]['otd']),
                reverse=rev
            )
        elif key == 'sort_period' and direction in ('asc','desc'):
            rev = (direction == 'desc')
            vacations_by_user_items.sort(
                key=lambda item: (
                    item[1]['vacation_start_dates'][0][0]
                    if item[1]['vacation_start_dates'] else dt.date.max
                ),
                reverse=rev
            )

    vacations_by_user = dict(vacations_by_user_items)
    
    return render(
        request,
        'vac_all.html',
        {
            'today': today,
            'year': year,
            'otd': otd,
            'len_vacations': len(filtered_vacations),
            'vacations_by_user': vacations_by_user,
            'holidays': holidays.get(year, {}),
            'otds_for_choise': otds_for_choise,
            'vacation_start_dates': vacation_start_dates,
            'show_button': True,
            'navbar_style': 'custom-navbar',
            'users_colors': users_colors,
            'bosses': list(bosses.keys()),
            'sort_otd': sort_otd,
            'sort_period': sort_period,
        }
    )


def vac_calendars(request, otd, year=None):
    current_year = dt.datetime.now().year
    future_years = [current_year + 1, current_year + 2]
    current_user_name = request.user.get_full_name()
    today = dt.datetime.today().date()
    bosses = get_bosses_dict()

    year = request.GET.get('year', current_year)
    year = int(year)

    years_in_vacations = set(Vacation.objects.values_list('year', flat=True))  # Исключаем дубликаты
    years_in_vacations.update([str(y) for y in future_years])
    years_range = sorted([int(y) for y in years_in_vacations], reverse=True)

    # Проверка, является ли пользователь боссом
    if current_user_name not in bosses.keys():
        otd_id = User_info.objects.filter(user_id=request.user.id).first().otd_number_id
        otd_users = User_info.objects.filter(otd_number_id=otd_id)
        otd_users_id = [user.user_id for user in otd_users]
        unit = Unit.objects.filter(id=otd_id).first()
        if unit:
            linked_units = [unit.description]
        else:
            linked_units = ["Неизвестный отдел"]
    else:
        if otd == 0:
            linked_units = bosses[current_user_name]
            otd_ids = [Unit.objects.filter(description=desc).first().id for desc in linked_units]
        else:
            otd_id = Unit.objects.filter(description=otd).first().id
            linked_units = [otd]
            otd_ids = [otd_id]

        otd_users = User_info.objects.filter(otd_number_id__in=otd_ids)
        otd_users_id = [user.user_id for user in otd_users]

    otds_for_choise = linked_units

    otd_data = []
    units = Unit.objects.filter(description__in=linked_units)
    for unit in units:
        unit_users = User_info.objects.filter(otd_number_id=unit.id)
        unit_user_ids = [user.user_id for user in unit_users]
        vacations_count = Vacation.objects.filter(user_id__in=unit_user_ids, year=str(year)).count()

        if vacations_count > 0:
            otd_data.append({
                'otd': unit.title,
                'otd_description': unit.description,
                'employees': unit_users.count(),
                'vacations': vacations_count
            })

    filtered_years_vacations_count = {}
    for y in years_range:
        filtered_years_vacations_count[y] = Vacation.objects.filter(
            user_id__in=otd_users_id, year=str(y)
        ).count()

    has_vacations_in_linked_units = any(count > 0 for count in filtered_years_vacations_count.values())

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'otd_data': otd_data,
            'linked_units': linked_units,
        })

    holidays_loaded = {}
    for y in years_range:
        holidays_loaded[y] = Holiday.objects.filter(date__year=y).exists()
    
    return render(
        request,
        'vac_calendars.html',
        {   'today': today,
            'year': year,
            'current_year': current_year,
            'otd': otd,
            'years_vacations_count': filtered_years_vacations_count,
            'otds_for_choise': otds_for_choise,
            'bosses': list(bosses.keys()),
            'current_user_name': current_user_name,
            'navbar_style': 'custom-navbar',
            'show_button': True,
            'years_range': years_range,
            'has_vacations_in_linked_units': has_vacations_in_linked_units,
            'holidays_loaded': holidays_loaded,
        }
    )


def vac_2(request, year, otd):
    if year == 0:
        year = dt.datetime.today().year
    today = dt.datetime.today().date()
    month_all = full_year(year)

    holidays, special_work_days = get_calendar_data(year)
    holidays_for_year = holidays.get(year, {})
    bosses = get_bosses_dict()

    bosses_list = list(bosses.keys())
    current_user_name = request.user.get_full_name()

    current_year = dt.datetime.now().year
    future_years = [current_year + 1, current_year + 2]

    years_in_vacations = set(Vacation.objects.values_list('year', flat=True))
    years_in_vacations.update([str(y) for y in future_years])
    years_range = sorted([int(y) for y in years_in_vacations])

    if request.user.get_full_name() not in bosses.keys():
        otd_id = User_info.objects.filter(user_id=request.user.id)[0].otd_number_id
        unit = Unit.objects.filter(id=otd_id).first()
        if unit:
            otd = int(unit.description)
        else:
            otd = 0
        otd_users = User_info.objects.filter(otd_number_id=otd_id) 
        otd_users_id = [user.user_id for user in otd_users] 
        vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))
        filtered_users_id = [request.user.id]
        otds_for_choise = [otd]
    else:
        if otd == 0:
            boss_departments = bosses[request.user.get_full_name()]
            otd_ids = [Unit.objects.filter(description=str(descr))[0].id for descr in boss_departments]
            otd_users = User_info.objects.filter(otd_number_id__in=otd_ids)
            otd_users_id = [user.user_id for user in otd_users]
            vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))
            filtered_users_id = otd_users_id.copy()
        else:
            otd_id = Unit.objects.filter(description=otd)[0].id  
            otd_users = User_info.objects.filter(otd_number_id=otd_id)  
            otd_users_id = [user.user_id for user in otd_users]  
            vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))
            filtered_users_id = otd_users_id.copy()
        otds_for_choise = bosses[request.user.get_full_name()]


        otd_ids = [Unit.objects.filter(description=descr)[0].id for descr in bosses[request.user.get_full_name()]]

        otd_users = User_info.objects.filter(otd_number_id__in=otd_ids) 
        otd_users_id = [user.user_id for user in otd_users]


    vacation_start_dates = {}
    for vac in vacations:
        if vac.user_id not in vacation_start_dates:
            vacation_start_dates[vac.user_id] = []

        vacation_start_dates[vac.user_id].append(vac.day_start)

    users_colors = {}
    for vac in vacations:
        if vac.user_id not in users_colors:
            users_colors[vac.user_id] = next(color_cycle)

    cross_vacation_groups = get_cross_vacations(vacations, users_colors, month_num_str)

    cross_vacation_groups.sort(key=lambda group: min(item['vac'].day_start for item in group))

    vacations_by_user = {}
    users_otd = User_info.objects.all()
    for vac in vacations:
        if vac.user_id not in vacations_by_user:
            vacations_by_user[vac.user_id] = {
                'color': users_colors[vac.user_id],
                'dates': [],
                'sum': 0,
                'otd': '',
                'user_id': vac.user_id,
                'vacation_start_dates': [],
                'vacation_end_dates': [],
                'user': vac.user
            }
            for u in users_otd:
                if u.user_id == vac.user_id:
                    vacations_by_user[vac.user_id]['otd'] = u.otd_number
                    break
        
        start_date = vac.day_start
        end_date = vac.day_end
        days_count = vac.how_long

        vacations_by_user[vac.user_id]['vacation_start_dates'].append((start_date, days_count))
        vacations_by_user[vac.user_id]['dates'].append(
            {'d': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} - {vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
            'vac_id': vac.id}
        )
        vacations_by_user[vac.user_id]['sum'] += vac.how_long
        vacations_by_user[vac.user_id]['vacation_end_dates'].append(end_date)

    for month, weeks in month_all.items():
        for week, days_in_week in weeks.items():
            for i in range(len(days_in_week)):
                data = {}
                date = ""
                if str(days_in_week[i]) != "":
                    month_number = get_key_from_dict_by_value(month_num_str, month)
                    day = today.replace(year=int(year), month=month_number, day=days_in_week[i])
                    for vac in vacations:
                        if day >= vac.day_start and day <= vac.day_end:
                            data[vac.user_id] = {
                                'date': f"{vac.day_start} - {vac.day_end}",
                                'color': users_colors[vac.user_id],
                                'vac_id': vac.id,
                            }
                    m = [k for k, v in month_num_str.items() if v == month][0]
                    date = today.replace(year=int(year), month=m, day=int(days_in_week[i]))
                month_all[month][week][i] = {'name': days_in_week[i], 'data': data, 'date': date}

    employees = User_info.objects.filter(
        user_id__in=filtered_users_id
    ).select_related('user', 'position')
    
    month_all_for_js = copy_dict_for_js(month_all)

    all_vac_for_js = {}
    for vac in vacations:
        all_vac_for_js[vac.id] = [str(vac.day_start), str(vac.day_end), vac.how_long]

    user_names = {user.id: user.get_full_name() for user in User.objects.filter(id__in=otd_users_id)}

    return render(
        request,
        'vac_new_calendar.html',
        {   
            'today': today,
            'year': year,
            'otd': otd,
            **month_all,
            'json_data': json.dumps(month_all_for_js),
            'json_data_vacs': json.dumps(all_vac_for_js),
            'user_names': json.dumps(user_names),
            'years_range': years_range,
            'cross_vacation_groups': cross_vacation_groups,
            'len_cross_vacations': len(cross_vacation_groups),
            'len_vacations': vacations.count(),
            'special_work_days': special_work_days,
            'holidays_json': json.dumps(holidays_for_year),
            'vacations_by_user': vacations_by_user,
            'otds_for_choise': otds_for_choise,
            'bosses': [key for key in bosses.keys()],
            'vacation_start_dates': vacation_start_dates,
            'current_user_id': request.user.id,
            'show_button': True,
            'show_add_leave_button': True,
            'bosses_list': json.dumps(bosses_list),
            'current_user_name': current_user_name,
            'navbar_style': 'custom-navbar',
            'show_vacation_link': True,
            'employees': employees,
        }
    )


def vacation_new(request, year):
    holidays_map, _ = get_calendar_data(year)
    current_holidays = holidays_map.get(year, {})
    bosses = get_bosses_dict()
    # Преобразуем holidays_map в множество date для расчёта
    # предполагаем, что holidays_map[year] — список date
    holidays_set = set(current_holidays)

    employee_id = request.GET.get('employee', None)

    current_user_name = request.user.get_full_name()
    is_boss = current_user_name in bosses

    if is_boss and employee_id:
        try:
            vacation = Vacation(user_id=int(employee_id))
        except (ValueError, User.DoesNotExist):
            vacation = Vacation(user_id=request.user.id)
    else:
        vacation = Vacation(user_id=request.user.id)

    form = VacationForm(request.POST or None, files=request.FILES or None, instance=vacation)

    employees = None
    departments = Unit.objects.all()

    selected_department = request.GET.get('department', None)

    if is_boss:
        boss_depts = bosses[current_user_name]
        dept_ids = Unit.objects.filter(description__in=boss_depts).values_list('id', flat=True)
        employees = User_info.objects.filter(otd_number_id__in=dept_ids).select_related('user', 'position')
        if selected_department and selected_department != "0":
            employees = employees.filter(otd_number__description=selected_department)

    selected_employee_name = current_user_name

    if form.is_valid():
        vac = form.save(commit=False)

        # Назначаем пользователя
        if is_boss and 'employee' in request.POST:
            vac.user_id = request.POST['employee']
            selected_employee_name = User.objects.get(id=vac.user_id).get_full_name()
        else:
            vac.user_id = request.user.id

        # Расчёт day_end и how_long
        if vac.day_start and vac.how_long and not vac.day_end:
            vac.day_end = calculate_end_date(vac.day_start, vac.how_long, holidays_set)
        elif vac.day_start and vac.day_end and not vac.how_long:
            vac.how_long = calculate_working_days(vac.day_start, vac.day_end, holidays_set)

        vac.year = vac.day_start.year
        vac.save()

        # Редиректы
        if employee_id:
            return redirect(f"{reverse('vac_all_vacations')}?user={vac.user_id}")
        elif is_boss and 'employee' in request.POST and str(vac.user_id) != str(request.user.id):
            return redirect(f"{reverse('vac_all_vacations')}?user={vac.user_id}")
        else:
            return redirect('vac_my_vacations')

    boss_depts = bosses.get(current_user_name, [])
    len_boss_departments = len(boss_depts)
    
    context = {
        'form': form,
        'user': request.user,
        'employees': employees,
        'is_boss': is_boss,
        'navbar_style': 'custom-navbar',
        'bosses': list(bosses.keys()),
        'year': year,
        'show_person': True,
        'holidays_json': json.dumps(current_holidays),
        'departments': departments,
        'len_boss_departments': len_boss_departments,
        'selected_department': selected_department,
        'selected_employee_id': employee_id,
    }
    return render(request, 'vacation_new.html', context)


def vacation_edit(request, year, vac_id):
    holidays_map, _ = get_calendar_data(year)
    current_holidays = holidays_map.get(year, {})
    holidays_set = set(current_holidays)
    bosses = get_bosses_dict()

    vac = get_object_or_404(Vacation, id=vac_id)
    form = VacationForm(request.POST or None, files=request.FILES or None, instance=vac)

    # Предзаполнение how_long
    if vac.day_start and vac.day_end:
        vac.how_long = calculate_working_days(vac.day_start, vac.day_end, holidays_set)

    employee_name = vac.user.get_full_name()
    employee_position = None
    if hasattr(vac.user, 'user_info'):
        user_info = vac.user.user_info.first()
        if user_info and user_info.position:
            employee_position = user_info.position.position

    if form.is_valid():
        vac = form.save(commit=False)

        if vac.day_start and vac.how_long and not vac.day_end:
            vac.day_end = calculate_end_date(vac.day_start, vac.how_long, holidays_set)
        elif vac.day_start and vac.day_end and not vac.how_long:
            vac.how_long = calculate_working_days(vac.day_start, vac.day_end, holidays_set)

        # Валидация: начало должно быть раньше конца
        if vac.day_start >= vac.day_end:
            form.add_error('day_end', 'Дата окончания должна быть позже даты начала')
        else:
            vac.year = vac.day_start.year
            vac.save()

            from_page = request.GET.get('from')
            if from_page == 'calendars':
                return redirect('vac_2', year=year, otd=0)
            elif from_page == 'all_vacations':
                return redirect(f"{reverse('vac_all_vacations')}?user={employee_name}")
            elif from_page == 'vac_all':
                return redirect('vac_all', otd=0)
            else:
                return redirect('vac_my_vacations')

    context = {
        'form': form,
        'user': request.user,
        'edit': True,
        'vac_id': vac_id,
        'year': year,
        'value': vac,
        'employee_name': employee_name,
        'employee_position': employee_position,
        'navbar_style': 'custom-navbar',
        'bosses': list(bosses.keys()),
        'redact_vac': True,
        'show_button': True,
        'holidays_json': json.dumps(current_holidays),
    }
    return render(request, 'vacation_edit.html', context)


def vacation_delete(request, vac_id):
    redirect_from = request.GET.get('from', None)

    vac = get_object_or_404(Vacation, id=vac_id)
    year = vac.day_start.year
    vac.delete()

    if redirect_from == 'calendars':
        return redirect('vac_2', year=year, otd=0)
    elif redirect_from == 'all_vacations':
        return redirect('vac_all_vacations')
    elif redirect_from == 'vac_all':
        return redirect('vac_all', otd=0)
    else:
        return redirect('vac_my_vacations')


def vacation_detail(request, vac_id):
    current_user_name = request.user.get_full_name()
    bosses = get_bosses_dict()

    # Проверка прав доступа
    if current_user_name not in bosses:
        vacation = get_object_or_404(Vacation, id=vac_id, user_id=request.user.id)
    else:
        boss_departments = bosses[current_user_name]
        department_ids = Unit.objects.filter(description__in=boss_departments).values_list('id', flat=True)
        allowed_users = User_info.objects.filter(otd_number_id__in=department_ids).values_list('user_id', flat=True)
        vacation = get_object_or_404(Vacation, id=vac_id, user_id__in=allowed_users)

    from_page = request.GET.get('from', None)
    context = {
        'start_date': vacation.day_start.strftime('%d.%m.%Y'),
        'end_date': vacation.day_end.strftime('%d.%m.%Y'),
        'days_count': vacation.how_long,
        'vacation_user_name': vacation.user.get_full_name(),
        'show_button': True,
        'vacation': vacation,
        'navbar_style': 'custom-navbar',
        'bosses': list(bosses.keys()),
        'show_vacation_detail': True,
        'from_page': from_page,
        'year': vacation.day_start.year,
    }
    return render(request, 'vacation_detail.html', context)


def vac_all_vacations(request):
    today = dt.datetime.today()
    bosses = get_bosses_dict()
    years_in_vacations = set(Vacation.objects.values_list('year', flat=True))
    year_range = sorted([int(y) for y in years_in_vacations], reverse=True)

    selected_otd = request.GET.get('otd', '') 
    selected_user = request.GET.get('user', '')
    selected_year = request.GET.get('year', '')

    current_user_name = request.user.get_full_name()
    filters = {}
    user_colors = {}

    # Ограничиваем по отделам, если босс
    if current_user_name in bosses:
        boss_departments = bosses[current_user_name]
        department_ids = Unit.objects.filter(
            description__in=boss_departments
        ).values_list('id', flat=True)
        filters['user__user_info__otd_number_id__in'] = department_ids

    if selected_otd:
        otd = Unit.objects.filter(title=selected_otd).first()
        if otd:
            filters['user__user_info__otd_number_id'] = otd.id

    if selected_user:
        try:
            uid = int(selected_user)
            filters['user__id'] = uid
        except ValueError:
            pass

    if selected_year:
        filters['year'] = selected_year

    vacations = Vacation.objects.filter(**filters)

    vacation_count = vacations.count()

    # Формирование списка отделов и пользователей для фильтров
    otds_for_choise = Unit.objects.all()
    if current_user_name in bosses:
        otds_for_choise = otds_for_choise.filter(
            description__in=bosses[current_user_name]
        )
        users_for_filter = User.objects.filter(
            user_info__otd_number_id__in=department_ids
        )
    else:
        users_for_filter = User.objects.all()

    years_vacations_count = {}
    for y in year_range:
        yf = {**filters, 'year': str(y)}
        years_vacations_count[y] = Vacation.objects.filter(**yf).count()

    # Собираем итоговый список отпусков
    vacations_list = []
    for vac in vacations:
        start_date = vac.day_start
        end_date = vac.day_end
        days_count = vac.how_long

        # Информация по пользователю
        user_info = vac.user.user_info.first()
        department = (user_info.otd_number.title 
                      if user_info and user_info.otd_number 
                      else 'Не указан')

        uid = vac.user.id
        if uid not in user_colors:
            user_colors[uid] = next(color_cycle)

        vacations_list.append({
            'user': vac.user.get_full_name(),
            'period': f"{start_date.day} {months_ru[start_date.month]} {start_date.year} - "
                      f"{end_date.day} {months_ru[end_date.month]} {end_date.year}",
            'days_count': days_count,
            'user_id': vac.user.id,
            'id': vac.id,
            'is_current': start_date <= today.date() <= end_date,
            'department': department,
            'color': user_colors[uid],
        })

    return render(
        request,
        'vac_all_vacations.html',
        {
            'today': today,
            'year': selected_year,
            'otds_for_choise': otds_for_choise,
            'users_for_filter': users_for_filter,
            'vacations': vacations_list,
            'selected_otd': selected_otd,
            'selected_user': selected_user,
            'selected_year': selected_year,
            'year_range': year_range,
            'years_vacations_count': years_vacations_count,
            'show_button': True,
            'navbar_style': 'custom-navbar',
            'bosses': list(bosses.keys()),
            'show_all_vacations': True,
            'vacation_count': vacation_count,
        }
    )


def vac_my_vacations(request):
    # Отображение отпусков текущего пользователя
    today = dt.datetime.today().date()
    year = today.year
    bosses = get_bosses_dict()

    # Получаем все отпуска текущего пользователя, отсортированные по убыванию даты начала
    vacations = Vacation.objects.filter(
        user_id=request.user.id
    ).order_by('-day_start')

    vacations_list = []
    for vac in vacations:
        start_date = vac.day_start
        end_date = vac.day_end
        days_count = vac.how_long

        vacations_list.append({
            'period': f"{start_date.day} {months_ru[start_date.month]} {start_date.year} - "
                      f"{end_date.day} {months_ru[end_date.month]} {end_date.year}",
            'days_count': days_count,
            'id': vac.id,
            'vacation_year': vac.day_start.year,
            'is_current': start_date <= today <= end_date,
        })

    # Определяем год для отображения
    vacation_year = year
    if vacations_list:
        try:
            vacation_year = int(vacations_list[0]['period'].split()[-1])
        except ValueError:
            pass

    return render(
        request,
        'vac_my_vacations.html',
        {
            'today': today,
            'year': year,
            'vacation_year': vacation_year,
            'vacations_list': vacations_list,
            'show_button': True,
            'navbar_style': 'custom-navbar',
            'bosses': list(bosses.keys()),
            'show_my_vacations': True,
        }
    )
        

def import_vacations(request):
    bosses = get_bosses_dict()
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            msg.error(request, "Файл не выбран.")
            return redirect(reverse('import_vacations'))

        if not file.name.endswith(('.xlsx', '.xls')):
            msg.error(request, "Неверный формат файла. Загрузите файл в формате .xlsx или .xls.")
            return redirect(reverse('import_vacations'))
        
        try:
            df = pd.read_excel(file)

            required_columns = ['Сотрудник', 'Отдел', 'Дата начала', 'Дата окончания']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                msg.error(request, f"В файле отсутствуют обязательные колонки: {', '.join(missing_columns)}.")
                return redirect(reverse('import_vacations'))

            df['Дата начала'] = pd.to_datetime(df['Дата начала'], format='%d.%m.%Y', dayfirst=True)
            df['Дата окончания'] = pd.to_datetime(df['Дата окончания'], format='%d.%m.%Y', dayfirst=True)

            if df['Дата начала'].isnull().any() or df['Дата окончания'].isnull().any():
                msg.error(request, "Некорректный формат дат в файле.")
                return redirect(reverse('import_vacations'))
            
            new_vacations = 0
            first_year = None 
            errors_found = False
            
            for index, row in df.iterrows():
                try:
                    full_name = row['Сотрудник'].strip()
                    parts = full_name.split()
                    if len(parts) == 2:
                        last_name, first_name = parts
                        patronymic = None
                    elif len(parts) == 3:
                        last_name, first_name, patronymic = parts
                    else:
                        last_name = parts[0]
                        first_name = parts[1]
                        patronymic = " ".join(parts[2:])

                    unit = Unit.objects.filter(description=row['Отдел']).first()
                    if not unit:
                        msg.error(
                            request,
                            f"Ошибка на строке {index + 1}: отдел '{row['Отдел']}' не найден."
                        )
                        errors_found = True
                        continue

                    qs = User_info.objects.filter(
                        user__first_name=first_name,
                        user__last_name=last_name,
                        otd_number=unit
                    )
                    if patronymic:
                        qs = qs.filter(user__patronymic=patronymic)
                    
                    user_info = qs.first()
                    if not user_info:
                        msg.error(
                            request,
                            f"Ошибка на строке {index + 1}: сотрудник '{full_name}' не принадлежит отделу '{row['Отдел']}'."
                        )
                        errors_found = True
                        continue

                    user = user_info.user

                    start_date = pd.to_datetime(row['Дата начала'], dayfirst=True)
                    end_date = pd.to_datetime(row['Дата окончания'], dayfirst=True)

                    if start_date > end_date:
                        msg.error(
                            request, 
                            f"Ошибка на строке {index + 1}: дата начала ({start_date.strftime('%d.%m.%Y')}) больше даты окончания ({end_date.strftime('%d.%m.%Y')})."
                        )
                        errors_found = True
                        continue
                    
                    if first_year is None:
                        first_year = start_date.year

                    existing_vacation = Vacation.objects.filter(
                        user=user,
                        day_start=start_date,
                        day_end=end_date,
                        user__user_info__otd_number=unit
                    ).exists()

                    if existing_vacation:
                        continue
                    
                    Vacation.objects.create(
                        user=user,
                        day_start=start_date,
                        day_end=end_date
                    )
                    new_vacations += 1
                except Exception:
                    errors_found = True
                    continue
            
            if not errors_found:
                msg.success(request, f"Добавлено новых отпусков: {new_vacations}.")
                return redirect(reverse('vac_2', kwargs={'otd': 0, 'year': first_year}))
            
        except Exception:
            msg.error(request, f"Ни одного отпуска не добавлено. Убедитесь, что все сотрудники уже добавлены.")
            return redirect(reverse('import_vacations'))

    context = {
        'navbar_style': 'custom-navbar',
        'import_vac': True,
        'show_button': True,
        'bosses': list(bosses.keys()),
    }

    return render(request, 'import_vacations.html', context)


def profile_view(request, user_id):
    bosses = get_bosses_dict()
    target_user = get_object_or_404(User, pk=user_id)
    user_info = get_object_or_404(User_info, user=target_user)
    today = dt.datetime.today().date()
    year = today.year

    vacations_qs = Vacation.objects.filter(
        user=target_user,
        day_end__gte=today
    ).order_by('day_start')

    vacations_list = []

    for vac in vacations_qs:
        start = vac.day_start
        end   = vac.day_end
        vacations_list.append({
            'period': (
                f"{start.day} {months_ru[start.month]} {start.year} – "
                f"{end.day} {months_ru[end.month]} {end.year}"
            ),
            'days_count': vac.how_long,
            'id': vac.id,
            'is_current': start <= today <= end,
            'vacation_year': start.year,
        })
    
    vacation_year = vacations_list[0]['vacation_year'] if vacations_list else None

    return render(
        request,
        'profile.html',
        {
            'today': today,
            'year': year,
            'vacation_year': vacation_year,
            'vacations_list': vacations_list,
            'navbar_style': 'custom-navbar',
            'bosses': list(bosses.keys()),
            'my_profile': True,
            'user_info': user_info,
            'is_own_profile':   (target_user == request.user),
        }
    )


def profile_edit(request, user_id):
    bosses = get_bosses_dict()
    target_user = get_object_or_404(User, pk=user_id)
    user_info   = get_object_or_404(User_info, user=target_user)
    departments = Unit.objects.all()
    all_tags    = Tag.objects.all().order_by('name')

    if request.method == "POST":
        action = request.POST.get('action')
        
        if action == 'archive':
            user_info.vacs_archiv = True
            user_info.save()
            return redirect(f"{reverse('employees')}?tab=employees")
        
        new_last  = request.POST.get("last_name", "").strip()
        new_first = request.POST.get("first_name", "").strip()
        new_pat   = request.POST.get("patronymic", "").strip()
        new_email = request.POST.get("email", "").strip()

        updated = False

        if target_user.first_name != new_first:
            target_user.first_name = new_first; updated = True
        if target_user.last_name  != new_last:
            target_user.last_name  = new_last;  updated = True
        if getattr(target_user, "patronymic", "") != new_pat:
            setattr(target_user, "patronymic", new_pat); updated = True
        if target_user.email != new_email:
            target_user.email = new_email; updated = True

        if updated:
            target_user.save()

        pos_name = request.POST.get("position", "").strip()
        if pos_name and (not user_info.position or user_info.position.position != pos_name):
            position, _ = Position.objects.get_or_create(position=pos_name)
            user_info.position = position
            updated = True

        dept_id = request.POST.get("department")
        new_dept = Unit.objects.get(id=dept_id) if dept_id else None
        if user_info.otd_number != new_dept:
            user_info.otd_number = new_dept
            updated = True

        tag_ids = request.POST.getlist("tags")
        if tag_ids is not None:
            valid_tags = Tag.objects.filter(id__in=tag_ids)
            user_info.tags.set(valid_tags)
            updated = True
        
        if updated:
            user_info.save()
            msg.success(request, "Сотрудник успешно обновлён.")

        return redirect('profile', user_id=target_user.id)

    return render(request, 'profile_edit.html', {
        'user_info': user_info,
        'departments': departments,
        'navbar_style': 'custom-navbar',
        'bosses': list(bosses.keys()),
        'profile_edit': True,
        'show_button': True,
        'is_own_profile': (target_user == request.user),
        'all_tags': all_tags, 
    })


def employees(request):
    bosses = get_bosses_dict()
    active_tab = request.GET.get('tab', 'employees')

    if request.method == 'POST':
        action = request.POST.get('action')
        ids = request.POST.getlist('ids') or request.POST.getlist('archive_ids')
        
        ids = [uid for uid in ids if uid != str(request.user.id)]
        
        if ids and action:
            qs = User_info.objects.filter(user_id__in=ids)
            if action == 'archive':
                qs.update(vacs_archiv=True)
            elif action == 'restore':
                qs.update(vacs_archiv=False)
            elif action == 'delete':
                # Удаляем сам объект User_info и пользователя
                users = [ui.user for ui in qs.select_related('user')]
                qs.delete()
                for user in users:
                    user.delete()
        
        if action == 'delete':
            tab = 'archive'
        else:
            tab = 'employees'
        return redirect(f"{reverse('employees')}?tab={tab}")

    dept_id = request.GET.get('dept', None)
    tag_name = request.GET.get('tag', None)

    try:
        current_user_info = request.user.user_info.first()
        current_dept = current_user_info.otd_number
        current_dept_id = current_dept.id if current_dept else None
    except Exception:
        current_dept_id = None
    
    selected_dept_title = None
    dept_obj = None
    if dept_id:
        try:
            dept_obj = Unit.objects.get(id=dept_id)
            selected_dept_title = dept_obj.title
        except ObjectDoesNotExist:
            dept_obj = None
    
    selected_tag_name = None
    if tag_name:
        tag_q = Tag.objects.filter(name=tag_name).first()
        if tag_q:
            selected_tag_name = tag_q.name
        else:
            selected_tag_name = None
    
    user_infos = (
        User_info.objects
        .filter(vacs_archiv=False)
        .select_related('user', 'position', 'otd_number')
        .prefetch_related('tags')
        .order_by('user__last_name', 'user__first_name')
    )

    if dept_obj:
        user_infos = user_infos.filter(otd_number=dept_obj)

    if selected_tag_name:
        user_infos = user_infos.filter(tags__name=selected_tag_name)
    
    user_colors = {}
    employees_list = []
    for ui in user_infos:
        user = ui.user
        uid = user.id

        if uid not in user_colors:
            user_colors[uid] = next(color_cycle)
            
        employees_list.append({
            'id': user.id,
            'full_name': user.get_full_name(),
            'position': ui.position.position if ui.position else '',
            'department': ui.otd_number.title if ui.otd_number else '',
            'department_id': ui.otd_number.id if ui.otd_number else None,
            'tags': [tag.name for tag in ui.tags.all()],
            'color': user_colors[uid],
        })

    units = Unit.objects.select_related('boss').all().order_by('title')
    departments_list = []
    for u in units:
        count = User_info.objects.filter(otd_number=u, vacs_archiv=False).count()
        departments_list.append({
            'id': u.id,
            'title': u.title,
            'count': count,
            'boss': u.boss.get_full_name() if u.boss else '',
        })

    tags_list = []
    if active_tab == 'tags':
        tags_qs = Tag.objects.annotate(
            employees_count=Count(
                'tags_info',
                filter=Q(tags_info__vacs_archiv=False)
            )
        ).order_by('name')

        for tag in tags_qs:
            tags_list.append({
                'id': tag.id,
                'name': tag.name,
                'employees_count': tag.employees_count,
            })

    archived_list = []
    if active_tab == 'archive':
        archived_qs = (
            User_info.objects
            .filter(vacs_archiv=True)
            .select_related('user', 'position', 'otd_number')
            .prefetch_related('tags')
            .order_by('user__last_name', 'user__first_name')
        )
        for ui in archived_qs:
            user = ui.user
            archived_list.append({
                'id': user.id,
                'full_name': user.get_full_name(),
                'position': ui.position.position if ui.position else '',
                'department': ui.otd_number.title if ui.otd_number else '',
                'tags': [tag.name for tag in ui.tags.all()],
            })
    
    managers_data = []
    if active_tab == 'managers':
        sup_ids = (
            User_info.objects
            .filter(supervisor__isnull=False)
            .values_list('supervisor_id', flat=True)
            .distinct()
        )
        for sup_id in sup_ids:
            sup_user = User.objects.filter(id=sup_id).first()
            if not sup_user:
                continue

            subordinate_depts = (
                User_info.objects
                .filter(supervisor_id=sup_id, otd_number__isnull=False)
                .values_list('otd_number__title', flat=True)
                .distinct()
            )

            direct_boss_depts = (
                Unit.objects
                .filter(boss_id=sup_id)
                .values_list('title', flat=True)
                .distinct()
            )

            combined = set(subordinate_depts) | set(direct_boss_depts)
            dept_list = sorted(combined)

            managers_data.append({
                'id': sup_user.id,
                'full_name': sup_user.get_full_name(),
                'departments': dept_list,
            })
    
    return render(request, 'employees.html', {
        'navbar_style': 'custom-navbar',
        'bosses': list(bosses.keys()),
        'employees': True,
        'show_button': True,
        'active_tab': active_tab,
        'employees_data': employees_list,
        'departments_data': departments_list,
        'tags_data': tags_list,
        'archived_data': archived_list,
        'selected_dept_title': selected_dept_title,
        'current_dept_id': current_dept_id,
        'selected_tag_name': selected_tag_name,
        'managers_data': managers_data,
    })


def get_department_employees(request):
    department_title = request.GET.get("department")
    if department_title:
        otd = Unit.objects.filter(title=department_title).first()
        if otd:
            employees = User.objects.filter(user_info__otd_number=otd)
        else:
            employees = User.objects.none()
    else:
        # Если отдел не выбран, возвращаем всех сотрудников
        employees = User.objects.all()
    
    employee_list = [
        {"id": emp.id, "full_name": emp.get_full_name()} for emp in employees
    ]
    return JsonResponse(employee_list, safe=False)


def get_initials(user):
    """
    Форматирует ФИО пользователя в виде "И.О. Фамилия".
    Предполагается, что у объекта user есть поля: first_name, patronymic, last_name.
    Если отчество состоит из нескольких частей (через дефис), каждая часть сокращается до первой буквы с точкой.
    """
    first = user.first_name.strip() if user.first_name else ""
    patronymic = user.patronymic.strip() if hasattr(user, 'patronymic') and user.patronymic else ""
    last = user.last_name.strip() if user.last_name else ""
    
    initial_first = first[0] + "." if first else ""
    
    if patronymic:
        parts = patronymic.split('-')
        initial_patronymic = "-".join([p.strip()[0] + "." for p in parts if p.strip()])
    else:
        initial_patronymic = ""
    
    initials = f"{initial_first}{initial_patronymic}".strip()
    
    result = f"{initials} {last}".strip() if last else initials
    return result


def export_vacations(request, year, otd):
    if otd == 0:
        msg.error(request, "Пожалуйста, выберите отдел для экспорта отпусков.")
        return redirect('vac_2', year=year, otd=otd)

    template_path = os.path.join(settings.BASE_DIR, 'export_templates', 'vacation_template.xlsx')
    if not os.path.exists(template_path):
        return HttpResponse("Шаблон не найден", status=404)

    wb = load_workbook(template_path)
    ws = wb.active

    # Получаем объект отдела по описанию
    otd_obj = Unit.objects.filter(description=otd).first()
    if not otd_obj:
        return HttpResponse("Такой отдел не найден", status=404)
    
    # Преобразуем title отдела в родительный падеж
    morph = pymorphy2.MorphAnalyzer()
    title = otd_obj.title
    match = re.match(r"(.+?)(\s*\(.*\))?$", title)
    if match:
        text_part = match.group(1).strip()
        num_part = match.group(2) or ""
    else:
        text_part = title
        num_part = ""
        
    words = text_part.split()
    inflected_words = []
    for word in words:
        parsed = morph.parse(word)[0]
        inflected = parsed.inflect({'gent'})
        if inflected:
            inflected_words.append(inflected.word)
        else:
            inflected_words.append(word)
    title_gen = " ".join(inflected_words) + num_part

    ws['A2'] = f"График отпусков на {year} год {title_gen}"

    # Получаем пользователей отдела
    otd_users = User_info.objects.filter(otd_number=otd_obj)
    user_ids = [user.user_id for user in otd_users]
    vacations = Vacation.objects.filter(user_id__in=user_ids, year=str(year))

    start_employee_row = 5

    template_styles = []
    for col in range(1, 6):
        cell = ws.cell(row=start_employee_row, column=col)
        template_styles.append({
            "font": copy.copy(cell.font),
            "border": copy.copy(cell.border),
            "fill": copy.copy(cell.fill),
            "number_format": copy.copy(cell.number_format),
            "protection": copy.copy(cell.protection),
            "alignment": copy.copy(cell.alignment),
        })

    template_row_height = ws.row_dimensions[start_employee_row].height

    for index, vac in enumerate(vacations):
        current_row = start_employee_row + index
        ws.row_dimensions[current_row].height = template_row_height

        if index > 0:
            for col in range(1, 6):
                new_cell = ws.cell(row=current_row, column=col)
                style = template_styles[col - 1]
                new_cell.font = style["font"]
                new_cell.border = style["border"]
                new_cell.fill = style["fill"]
                new_cell.number_format = style["number_format"]
                new_cell.protection = style["protection"]
                new_cell.alignment = style["alignment"]

        ws.cell(row=current_row, column=1).value = index + 1
        ws.cell(row=current_row, column=2).value = vac.user.get_full_name()
        try:
            position = vac.user.user_info.first().position.position
        except Exception:
            position = ""
        ws.cell(row=current_row, column=3).value = position
        ws.cell(row=current_row, column=4).value = vac.day_start.strftime('%d.%m.%Y')
        ws.cell(row=current_row, column=5).value = vac.how_long

    filename = f"График отпусков {otd} {year}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

    #Формирование нижнего блока
    footer_start_row = start_employee_row + len(vacations) + 2

    custom_font = Font(name="Times New Roman", size=14)
    thin_border = Border(bottom=Side(style='thin'))

    ws.merge_cells(start_row=footer_start_row, start_column=1, end_row=footer_start_row, end_column=2)
    cell = ws.cell(row=footer_start_row, column=1)
    cell.value = "СОГЛАСОВАНО:"
    cell.font = Font(name="Times New Roman", size=14)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    boss = otd_obj.boss
    boss_info = User_info.objects.filter(user=boss).first() if boss else None
    dept_manager_position = boss_info.position.position if boss_info and boss_info.position else ""
    dept_manager_full_name = boss.get_full_name() if boss else ""

    if boss_info and boss_info.supervisor:
        supervisor = boss_info.supervisor
        supervisor_info = User_info.objects.filter(user=supervisor).first()
        supervisor_position = supervisor_info.position.position if supervisor_info and supervisor_info.position else ""
        supervisor_full_name = supervisor.get_full_name()
    else:
        supervisor_position = ""
        supervisor_full_name = ""

    if supervisor_full_name:
        supervisor_row = footer_start_row + 1
        ws.row_dimensions[supervisor_row].height = 63

        ws.merge_cells(start_row=supervisor_row, start_column=1, end_row=supervisor_row, end_column=2)
        sup_cell1 = ws.cell(row=supervisor_row, column=1)
        sup_cell1.value = supervisor_position
        sup_cell1.font = custom_font
        sup_cell1.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        ws.merge_cells(start_row=supervisor_row, start_column=4, end_row=supervisor_row, end_column=5)
        sup_cell2 = ws.cell(row=supervisor_row, column=4)
        sup_cell2.value = get_initials(supervisor)
        sup_cell2.font = custom_font
        sup_cell2.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        cell_3 = ws.cell(row=supervisor_row, column=3)
        cell_3.border = thin_border
        cell_3.font = custom_font

        dept_manager_row = supervisor_row + 2
        ws.row_dimensions[dept_manager_row].height = 63

        ws.merge_cells(start_row=dept_manager_row, start_column=1, end_row=dept_manager_row, end_column=2)
        dm_cell1 = ws.cell(row=dept_manager_row, column=1)
        dm_cell1.value = dept_manager_position
        dm_cell1.font = custom_font
        dm_cell1.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        ws.merge_cells(start_row=dept_manager_row, start_column=4, end_row=dept_manager_row, end_column=5)
        dm_cell2 = ws.cell(row=dept_manager_row, column=4)
        dm_cell2.value = get_initials(boss)
        dm_cell2.font = custom_font
        dm_cell2.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        final_cell3 = ws.cell(row=dept_manager_row, column=3)
        final_cell3.border = thin_border
        final_cell3.font = custom_font

    else:
        dept_manager_row = footer_start_row + 1
        ws.row_dimensions[dept_manager_row].height = 63

        ws.merge_cells(start_row=dept_manager_row, start_column=1, end_row=dept_manager_row, end_column=2)
        dm_cell1 = ws.cell(row=dept_manager_row, column=1)
        dm_cell1.value = dept_manager_position
        dm_cell1.font = custom_font
        dm_cell1.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        ws.merge_cells(start_row=dept_manager_row, start_column=4, end_row=dept_manager_row, end_column=5)
        dm_cell2 = ws.cell(row=dept_manager_row, column=4)
        dm_cell2.value = get_initials(boss)
        dm_cell2.font = custom_font
        dm_cell2.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        cell_3 = ws.cell(row=dept_manager_row, column=3)
        cell_3.border = thin_border
        cell_3.font = custom_font

    wb.save(response)
    return response


def save_vacations(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            for vacation_data in data:
                # Создаем объект с валидацией
                vacation = Vacation(
                    user_id=vacation_data.get('employee'),
                    day_start=vacation_data.get('day_start'),
                    day_end=vacation_data.get('day_end')
                )               
                # Вызываем полную валидацию
                vacation.full_clean()
                vacation.save()
                
            return JsonResponse({'status': 'success'})
        except ValidationError as e:
            return JsonResponse({'status': 'error', 'message': dict(e)})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'invalid method'})


def tag_create(request):
    bosses = get_bosses_dict()
    form = TagForm(request.POST or None)

    context = {
        'form': form,
        'navbar_style': 'custom-navbar',
        'employees_add_tag': True,
        'bosses': list(bosses.keys()),
        'show_button': True,
    }
    
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect(f"{reverse('employees')}?tab=tags")
    
    return render(request, 'tag_form.html', context)


def tag_update(request, pk):
    bosses = get_bosses_dict()
    tag = get_object_or_404(Tag, pk=pk)

    form = TagForm(request.POST or None, instance=tag)

    context = {
        'form': form,
        'navbar_style': 'custom-navbar',
        'employees_edit_tag': True,
        'is_edit': True,
        'tag': tag,
        'bosses': list(bosses.keys()),
        'show_button': True,
    }
    
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('employees')}?tab=tags")

    return render(request, 'tag_form.html', context)


def tag_delete(request, pk):
    tag = get_object_or_404(Tag, pk=pk)
    if request.method == 'POST':
        tag.delete()
        return redirect(f"{reverse('employees')}?tab=tags")
    return redirect(f"{reverse('employees')}?tab=tags")


def unit_create(request):
    bosses = get_bosses_dict()
    form = UnitForm(request.POST or None)

    context = {
        'form': form,
        'navbar_style': 'custom-navbar',
        'employees_add_unit': True,
        'bosses': list(bosses.keys()),
        'show_button': True,
    }

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect(f"{reverse('employees')}?tab=departments")
    
    return render(request, 'unit_form.html', context)


def unit_update(request, pk):
    bosses = get_bosses_dict()
    unit = get_object_or_404(Unit, pk=pk)

    form = UnitForm(request.POST or None, instance=unit)
    
    context = {
        'form': form,
        'navbar_style': 'custom-navbar',
        'employees_edit_unit': True,
        'bosses': list(bosses.keys()),
        'show_button': True,
        'is_edit': True,
        'department': unit,
    }
    
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect(f"{reverse('employees')}?tab=departments")
    
    return render(request, 'unit_form.html', context)


def unit_delete(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        unit.delete()
        return redirect(f"{reverse('employees')}?tab=departments")
    return redirect(f"{reverse('employees')}?tab=departments")


def employee_create(request):
    bosses = get_bosses_dict()
    departments = Unit.objects.all().order_by('title')
    all_tags = Tag.objects.all().order_by('name')

    context = {
        'departments': departments,
        'all_tags': all_tags,
        'navbar_style': 'custom-navbar',
        'show_button': True,
        'employees_add': True,
        'bosses': list(bosses.keys()),
    }
    
    if request.method == "POST":
        last_name = request.POST.get("last_name", "").strip()
        first_name = request.POST.get("first_name", "").strip()
        patronymic = request.POST.get("patronymic", "").strip()
        email = request.POST.get("email", "").strip().lower()
        position_nm = request.POST.get("position", "").strip()
        dept_id = request.POST.get("department", "").strip()
        tag_ids = request.POST.getlist("tags")

        # Проверка обязательных полей
        if not (last_name and first_name and email):
            msg.error(request, "Пожалуйста, заполните все обязательные поля (ФИО и email).")
            return render(request, 'employee_create.html', context)

        # Уникальность email
        if User.objects.filter(email=email).exists():
            msg.error(request, "Пользователь с таким email уже существует.")
            return render(request, 'employee_create.html', context)

        username = email.split("@")[0]
        # Генерация пароля
        temp_password = User.objects.make_random_password()

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=temp_password,
                    first_name=first_name,
                    last_name=last_name,
                )
                if hasattr(user, 'patronymic'):
                    user.patronymic = patronymic
                    user.save()

                position = None
                if position_nm:
                    position, _ = Position.objects.get_or_create(position=position_nm)

                department = None
                if dept_id:
                    try:
                        department = Unit.objects.get(id=dept_id)
                    except ObjectDoesNotExist:
                        department = None

                ui = User_info.objects.create(
                    user=user,
                    otd_number=department,
                    position=position
                )

                if tag_ids:
                    for t_id in tag_ids:
                        try:
                            tag_obj = Tag.objects.get(id=t_id)
                            ui.tags.add(tag_obj)
                        except Tag.DoesNotExist:
                            continue

        except IntegrityError:
            msg.error(request, "Ошибка при создании пользователя. Попробуйте ещё раз.")
            return render(request, 'employee_create.html', context)

        full_name = " ".join(filter(None, [last_name, first_name, patronymic]))
        
        subject = "Ваши учётные данные от информационного ресурса"
        message = (
            f"Здравствуйте, {full_name}!\n\n"
            f"Ваш логин: {username}\n"
            f"Ваш пароль: {temp_password}\n\n"
            "Пожалуйста, смените пароль при первом входе.\n\n"
            "С уважением,\n"
            "Администратор ресурса."
        )
        send_mail(subject, message, None, [email], fail_silently=False)

        msg.success(request, f"Сотрудник {full_name} успешно создан.")
        return redirect(f"{reverse('employees')}?tab=employees")

    return render(request, 'employee_create.html', context)


def manager_create(request):
    bosses = get_bosses_dict()
    all_users = User.objects.filter(user_info__vacs_archiv=False).order_by('last_name', 'first_name')
    departments = Unit.objects.all().order_by('title')

    context = {
        'all_users': all_users,
        'departments': departments,
        'navbar_style': 'custom-navbar',
        'show_button': True,
        'manager_add': True,
        'bosses': list(bosses.keys()),
    }
    
    if request.method == 'POST':
        supervisor_id = request.POST.get('employee')
        selected_depts = request.POST.getlist('departments')

        if not supervisor_id or not selected_depts:
            return render(request, 'manager_form.html', context)

        supervisor_user = get_object_or_404(User, pk=supervisor_id)

        # Пройдём по каждому выбранному отделу
        for dep_id in selected_depts:
            unit = Unit.objects.filter(id=dep_id).first()
            if not unit:
                continue

            if not unit.boss:
                continue

            dept_head_user = unit.boss

            try:
                head_info = User_info.objects.get(user=dept_head_user)
            except User_info.DoesNotExist:
                continue

            if head_info.user_id == supervisor_user.id:
                continue

            head_info.supervisor = supervisor_user
            head_info.save()

        msg.success(
            request,
            f"{supervisor_user.get_full_name()} назначен(а) супервайзером для выбранных отделов."
        )
        return redirect(f"{reverse('employees')}?tab=managers")

    return render(request, 'manager_form.html', context)


def manager_edit(request, supervisor_id):
    bosses = get_bosses_dict()
    all_users = User.objects.filter(user_info__vacs_archiv=False).order_by('last_name', 'first_name')
    departments = Unit.objects.all().order_by('title')

    sup_user = get_object_or_404(User, pk=supervisor_id)
    initial_employee = sup_user.id

    subordinate_infos = User_info.objects.filter(supervisor=sup_user, otd_number__isnull=False)
    depts_from_subs = { ui.otd_number.id for ui in subordinate_infos }

    direct_boss_depts = { unit.id for unit in Unit.objects.filter(boss=sup_user) }

    initial_departments = sorted(depts_from_subs | direct_boss_depts)

    context = {
        'all_users': all_users,
        'departments': departments,
        'initial_employee': initial_employee,
        'initial_departments': initial_departments,
        'navbar_style': 'custom-navbar',
        'show_button': True,
        'manager_edit': True,
        'supervisor_id': supervisor_id,
        'bosses': list(bosses.keys()),
    }
    
    if request.method == 'POST':
        form_supervisor_id = request.POST.get('supervisor_id') or supervisor_id
        employee_id = request.POST.get('employee')
        selected_depts = request.POST.getlist('departments')

        if not employee_id or not selected_depts:
            return render(request, 'manager_form.html', context)

        supervisor_user = get_object_or_404(User, pk=employee_id)

        if form_supervisor_id:
            old_sup = get_object_or_404(User, pk=form_supervisor_id)
            User_info.objects.filter(supervisor=old_sup).update(supervisor=None)

        for dep_id in selected_depts:
            unit = Unit.objects.filter(id=dep_id).first()
            if not unit or not unit.boss:
                continue

            dept_head_user = unit.boss
            try:
                head_info = User_info.objects.get(user=dept_head_user)
            except User_info.DoesNotExist:
                continue

            if head_info.user_id == supervisor_user.id:
                continue

            head_info.supervisor = supervisor_user
            head_info.save()

        return redirect(f"{reverse('employees')}?tab=managers")
    
    return render(request, 'manager_form.html', context)